# -*- coding: UTF-8 -*-
import os
import openpyxl
from BaseLib.HandleFile import file_utils
from BaseLib.LoggerConfig import logger



class HandleExec(object):


    def __init__(self,file_path="data/data.xlsx"):
        self.file_path = file_utils.location_file(file_path)
        self.load_excel = self.__load_excel(self.file_path)



    def __load_excel(self,file_path):
        '''
        加载sheet
        '''

        if not os.path.exists(file_path):
            logger.info("文件路径不存在:{0}".format(file_path))
            raise FileNotFoundError("文件不存在:{}".format(file_path))
        return openpyxl.load_workbook(file_path)


    def get_sheet_data(self,index=None):
        '''
        加载所有sheet的内容
        '''
        sheet_name =  self.load_excel.sheetnames
        if index == None:
            index = 0
        data = self.load_excel[sheet_name[index]]
        return data

    def get_cell_value(self, row, cols):
        '''
        获取某一个单元格内容
        '''
        data = self.get_sheet_data().cell(row=row, column=cols).value
        return data

    def get_rows(self):
        '''
        获取行数
        '''
        row = self.get_sheet_data().max_row
        return row

    def get_data(self, row):
        '''
        获取某一行的内容
        '''
        row_item = []
        row_list= []
        count = 0
        for i in self.get_sheet_data()[row+1]:
            if i.value and count!=0:
                row_item.append(i.value)
            count=count+1
        row_list.append(row_item)
        return row_list

    '''
        根据第一列值来找到所有相同的行
    '''
    def get_rows_values(self, value):
        row_list = []
        value_list=[]
        data_execl = self.get_sheet_data()
        for i in data_execl:
            if i[0].value == value:
                for item in range(len(i)) :
                    if isinstance(i[item].value,str) and item>0 and (i[item].value).replace(" ", ""):
                        row_list.append(i[item].value)
                value_list.append(row_list)
                row_list=[]
        return value_list




    def get_columns_value(self, key=None):
        '''
        获取某一列得数据
        '''
        columns_list = []
        if key == None:
            key = 'A'
        columns_list_data = self.get_sheet_data()[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list



    def get_rows_number(self, case_id):
        '''
        根据第一列数值获取行号
        '''
        num = 1
        cols_data = self.get_columns_value()
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num + 1
        return num

    def get_excel_data(self):
        '''
        获取excel里面所有的数据
        '''
        data_list = []
        for i in range(self.get_rows()):
            data_list.append(self.get_rows_value(i + 2))
        return data_list

    def excel_write_data(self, row, cols, value):
        '''
        写入数据
        '''
        wb =  self.load_excel
        wr = wb.active
        wr.cell(row, cols, value)
        wb.save(self.file_path)



handle_exec =HandleExec()

if __name__ == '__main__':

    print(handle_exec.get_rows_values("test_04_address_manage"))
    print(handle_exec.get_data(1))